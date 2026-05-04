import os
from copy import copy
from openpyxl import load_workbook


def fill_excel_template(template_path, data, output_dir):
    """
    Fill EnergyBae solar excel template
    Preserves formulas / formatting
    """

    wb = load_workbook(template_path)
    ws = wb.active

    # -----------------------------
    # Header fields (LEFT BLOCK)
    # -----------------------------
    ws["D1"] = data["consumer_name"]
    ws["D2"] = data["consumer_number"]
    ws["D3"] = float(data["fixed_charges"])
    ws["D4"] = data["sanctioned_load"]
    ws["D5"] = data["connection_type"]

    # -----------------------------
    # Monthly units
    # D9 → D20
    # -----------------------------
    start_row = 9

    monthly_units = data["monthly_units"]

    if len(monthly_units) < 12:
        monthly_units = [0] * (12 - len(monthly_units)) + monthly_units

    monthly_units = monthly_units[:12]

    for i, units in enumerate(monthly_units):
        ws[f"D{start_row+i}"] = units

    # -----------------------------
    # Bill amounts
    # Put latest amount in E20
    # Optionally replicate for history
    # -----------------------------
    ws["E20"] = float(data["latest_bill_amount"])

    # -----------------------------
    # Save
    # -----------------------------
    output_path = os.path.join(
        output_dir,
        "EnergyBae_Solar_Report.xlsx"
    )

    wb.save(output_path)

    return output_path